# Generating an editable, printable Excel calendar for 2026
# - Weeks start on Sunday
# - No holidays included
# - One worksheet per month (Jan–Dec)
# - Simple formatting for printing (landscape, fit to width)

import calendar
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.page import PageMargins

year = 2026
calendar.setfirstweekday(calendar.SUNDAY)  # Weeks start on Sunday
month_names = [calendar.month_name[i] for i in range(1,13)]
columns = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

# Create an Excel writer
output_path = "/mnt/data/Calendar_2026.xlsx"
with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    for month in range(1, 13):
        # monthcalendar returns weeks as lists of day numbers (0 for padding)
        weeks = calendar.monthcalendar(year, month)
        # Convert 0 -> "" for empty cells
        data = [[day if day != 0 else "" for day in week] for week in weeks]
        # Make DataFrame (rows = weeks, cols = Sun..Sat)
        df = pd.DataFrame(data, columns=columns)
        # Add a header row with the month and year by writing to the sheet title cell later
        sheet_name = f"{month:02d} - {month_names[month-1]}"
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)

    # Add a cover sheet (optional)
    cover_df = pd.DataFrame({"A": [f"Calendar - {year}", "", "Weeks start on Sunday", "No holidays included"]})
    cover_df.to_excel(writer, sheet_name="Cover", index=False, header=False)

# Post-process formatting with openpyxl
wb = load_workbook(output_path)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # Center align all cells and set font
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Make day numbers slightly larger
            if isinstance(cell.value, int):
                cell.font = Font(size=11, bold=False)
    # Set column widths to be more square-like
    for col in range(1, 8):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 14

    # Merge top row cells for month title if it's a month sheet
    if sheet_name != "Cover":
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        ws.cell(row=1, column=1).value = sheet_name.replace(" - ", " ")
        ws.cell(row=1, column=1).font = Font(size=14, bold=True)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

    # Page setup for printing: landscape, fit to width 1 page
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # allow multiple pages tall if needed
    # Reasonable margins for printing
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)

# Save workbook
wb.save(output_path)

# Display a preview table for January to the user
import caas_jupyter_tools as tools, pandas as pd
jan_sheet = wb["01 - January"]
# Read rows corresponding to the calendar area (including title row)
data_preview = []
for r in jan_sheet.iter_rows(min_row=1, max_row=jan_sheet.max_row, min_col=1, max_col=7, values_only=True):
    data_preview.append(list(r))
preview_df = pd.DataFrame(data_preview)
tools.display_dataframe_to_user("Preview - January 2026", preview_df)

print(f"Saved calendar to: {output_path}")
